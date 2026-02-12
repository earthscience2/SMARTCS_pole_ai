#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
4. merge_data 전체에 대해 2차 Hard 모델(13. evaluate_hard_model_2nd)의 검출 결과를
엑셀로 정리해서 make_ai/test 아래에 저장하는 스크립트.

[프로젝트/전주 불일치 원인]
  프로젝트·전주는 predictions.csv 의 csv_path 경로를 파싱해 정함 (4. merge_data/break|normal/프로젝트명/전주ID).
  따라서 4. merge_data 폴더 구조가 잘못되면(다른 프로젝트 폴더 아래 전주가 있으면) 엑셀에 잘못된 프로젝트가 찍힘.
  이 스크립트는 3-1 단계에서 DB(tb_anal_state, tb_pole)로 전주 소속 프로젝트를 조회해 경로와 다르면 보정함.

사용 방법(권장 순서):
1) 먼저 13. evaluate_hard_model_2nd.py 를 4. merge_data 전체에 대해 실행
   예) PowerShell:
       cd C:\Users\slhg1\OneDrive\Desktop\SMARTCS_Pole\make_ai
       python "13. evaluate_hard_model_2nd.py" --data-dir "4. merge_data" --no-test-only

   → 최신 2차 Hard 모델(run)을 사용해서
     13. evaluate_hard_model_2nd/<timestamp>/test/predictions.csv 가 생성됨.

2) 그 다음 이 스크립트를 실행
   예)
       python "14. export_2nd_eval_to_excel.py"

   → make_ai/test/hard_2nd_merge_data_predictions.xlsx 로 저장.
"""

import os
import sys
import glob
import json
from pathlib import Path
import importlib.util

import numpy as np
import pandas as pd
from tensorflow import keras
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def _get_project_for_pole_from_db(server: str, pole_id: str, project_root: Path):
    """DB에서 전주(pole_id)의 실제 소속 프로젝트(groupname) 조회. 실패 시 None."""
    if not pole_id or not server:
        return None
    try:
        if str(project_root) not in sys.path:
            sys.path.insert(0, str(project_root))
        from config import poledb as PDB
        PDB.poledb_init(server)
        if not hasattr(PDB, "poledb_conn") or PDB.poledb_conn is None:
            return None
        q = "SELECT groupname FROM tb_anal_state WHERE poleid = %s LIMIT 1"
        result = PDB.poledb_conn.do_select_pd(q, [pole_id])
        if result is not None and not result.empty:
            return str(result.iloc[0]["groupname"]).strip()
        q2 = "SELECT groupname FROM tb_pole WHERE poleid = %s LIMIT 1"
        result2 = PDB.poledb_conn.do_select_pd(q2, [pole_id])
        if result2 is not None and not result2.empty:
            return str(result2.iloc[0]["groupname"]).strip()
    except Exception:
        pass
    return None


def _get_latest_eval_dir(base: Path) -> Path:
    """13. evaluate_hard_model_2nd 아래에서 가장 최근 평가 run 디렉터리 반환."""
    if not base.exists():
        raise FileNotFoundError(f"평가 디렉터리를 찾을 수 없습니다: {base}")
    subs = [d for d in base.iterdir() if d.is_dir()]
    if not subs:
        raise FileNotFoundError(f"평가 run 디렉터리가 없습니다: {base}")
    return max(subs, key=lambda d: d.name)


def _get_latest_light_model_run(models_base: Path):
    """7. light_models 안에서 checkpoints/best.keras 가 있는 run 중 이름 기준 최신 폴더 반환."""
    if not models_base.exists():
        return None
    candidates = []
    for d in models_base.iterdir():
        if not d.is_dir():
            continue
        ckpt = d / "checkpoints" / "best.keras"
        if ckpt.exists():
            candidates.append((d.name, d, ckpt))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0]  # (name, dir, best_ckpt_path)


def _build_project_to_server_mapping(current_dir: Path):
    """
    anal2_poles_all_*.json (2. anal_pole_list)에서 프로젝트 -> 서버(main/is/kh) 매핑 생성.
    3.1. check_raw_pole_data_info.py 의 build_project_to_server_mapping 와 동일 개념.
    """
    anal_dir = current_dir / "2. anal_pole_list"
    pattern = str(anal_dir / "anal2_poles_all_*.json")
    files = glob.glob(pattern)
    if not files:
        print(f"경고: anal2_poles_all JSON을 찾을 수 없어 서버 매핑을 생성하지 못했습니다: {pattern}")
        return {}
    anal_file = max(files, key=os.path.getmtime)
    print(f"프로젝트-서버 매핑 로드: {anal_file}")
    with open(anal_file, "r", encoding="utf-8") as f:
        anal_data = json.load(f)
    servers_data = anal_data.get("servers", {})
    project_to_server = {}
    for server, server_info in servers_data.items():
        projects = server_info.get("projects", {})
        if isinstance(projects, dict):
            for proj_name in projects.keys():
                project_to_server[proj_name] = server
    return project_to_server


def _load_set_light_train_data_module(current_dir: Path):
    """6. set_light_train_data 모듈 로드 (Light 모델 전처리 재사용)."""
    script_path = current_dir / "6. set_light_train_data.py"
    if not script_path.exists():
        print(f"경고: 6. set_light_train_data.py 를 찾을 수 없습니다: {script_path}")
        return None
    spec = importlib.util.spec_from_file_location("set_light_train_data", script_path)
    mod = importlib.util.module_from_spec(spec)
    sys_modules = __import__("sys").modules
    sys_modules["set_light_train_data"] = mod
    assert spec.loader is not None
    spec.loader.exec_module(mod)
    return mod


def main():
    # weight_test 폴더 기준: 입력은 make_ai(상위), 출력은 weight_test
    current_dir = Path(__file__).resolve().parent  # weight_test
    make_ai_dir = current_dir.parent  # make_ai

    # 1) 가장 최근 2차 평가 run 찾기 (make_ai 내 13번)
    eval_base = make_ai_dir / "13. evaluate_hard_model_2nd"
    eval_dir = _get_latest_eval_dir(eval_base)
    # scope 자동 선택: test 우선, 없으면 all
    test_dir = eval_dir / "test"
    all_dir = eval_dir / "all"
    if (test_dir / "predictions.csv").exists():
        scope_dir = test_dir
    elif (all_dir / "predictions.csv").exists():
        scope_dir = all_dir
    else:
        scope_dir = test_dir  # 오류 메시지를 위해 기본값 유지
    csv_path = scope_dir / "predictions.csv"

    if not csv_path.exists():
        raise FileNotFoundError(f"predictions.csv 를 찾을 수 없습니다: {csv_path}\n"
                                f"※ 먼저 13. evaluate_hard_model_2nd.py 를 실행해 주세요.")

    print(f"최근 2차 평가 run 사용: {eval_dir.name}")
    print(f"읽는 파일: {csv_path}")

    df = pd.read_csv(csv_path, encoding="utf-8-sig")

    # 2) anal2_poles_all_* 기반 프로젝트 -> 서버(main/is/kh) 매핑 로드 (make_ai 내 2번)
    project_to_server = _build_project_to_server_mapping(make_ai_dir)

    # 3) csv_path에서 서버(main/is/kh), 프로젝트, 전주ID 추출
    def _parse_path_info(p: str):
        try:
            path = Path(p)
            parts = [str(part) for part in path.parts]
            server = project = pole_id = None
            if "4. merge_data" in parts:
                idx = parts.index("4. merge_data")
                # parts[idx+1] = break/normal, parts[idx+2] = 프로젝트명, parts[idx+3] = 전주ID
                if idx + 2 < len(parts):
                    project = parts[idx + 2]
                    server = project_to_server.get(project, None)
                if idx + 3 < len(parts):
                    pole_id = parts[idx + 3]
            return server, project, pole_id
        except Exception:
            return None, None, None

    servers, projects, pole_ids = [], [], []
    for p in df["csv_path"].astype(str):
        s, pj, pid = _parse_path_info(p)
        servers.append(s)
        projects.append(pj)
        pole_ids.append(pid)

    df["server"] = servers
    df["project"] = projects
    df["pole_id"] = pole_ids

    # 3-1) DB에서 전주 소속 프로젝트 확인 후 보정 (경로 파싱만으로는 폴더 잘못 시 프로젝트/전주 불일치 발생)
    project_root = make_ai_dir.parent
    seen_mismatch = set()
    for i in range(len(df)):
        pid = pole_ids[i] if i < len(pole_ids) else None
        srv = servers[i] if i < len(servers) else None
        path_project = projects[i] if i < len(projects) else None
        if not pid or not srv:
            continue
        db_project = _get_project_for_pole_from_db(srv, str(pid).strip(), project_root)
        if db_project is None:
            continue
        if path_project != db_project:
            key = (srv, pid, path_project, db_project)
            if key not in seen_mismatch:
                seen_mismatch.add(key)
                print(f"  [보정] 전주 {pid}: 경로상 프로젝트 '{path_project}' → DB 소속 '{db_project}' 로 수정")
            projects[i] = db_project
    df["project"] = projects

    # 4) label(0/1)을 사람이 보기 쉬운 텍스트로 추가
    df["label_text"] = df["label"].map({1: "break", 0: "normal"}).fillna("unknown")

    # 5) Light 모델(7. light_models 최신 best.keras) 로드 (make_ai 내 7번)
    light_models_base = make_ai_dir / "7. light_models"
    latest_light = _get_latest_light_model_run(light_models_base)
    light_prob_by_path = {}
    if latest_light is not None:
        run_name, run_dir, best_ckpt_path = latest_light
        print(f"Light 모델 run 사용: {run_name}")
        print(f"Light 모델 경로: {best_ckpt_path}")
        try:
            light_model = keras.models.load_model(str(best_ckpt_path), compile=False)
        except Exception as e:
            print(f"경고: Light 모델 로드 실패: {e}")
            light_model = None

        if light_model is not None:
            std_mod = _load_set_light_train_data_module(make_ai_dir)
            if std_mod is None:
                print("경고: 6. set_light_train_data 모듈을 로드하지 못해 Light 추론을 생략합니다.")
            else:
                prepare_sequence_from_csv = std_mod.prepare_sequence_from_csv
                resize_img_height = std_mod.resize_img_height

                unique_paths = sorted(set(df["csv_path"].astype(str)))
                imgs = []
                paths_for_pred = []
                target_h = 304
                print(f"Light 모델 추론용 샘플 수: {len(unique_paths)}")
                for p in unique_paths:
                    try:
                        r = prepare_sequence_from_csv(csv_path=p, sort_by="height", feature_min_max=None)
                    except Exception:
                        r = None
                    if r is None:
                        continue
                    img, meta = r
                    img = resize_img_height(img, target_h=target_h)
                    imgs.append(img)
                    paths_for_pred.append(p)

                if imgs:
                    X_light = np.array(imgs, dtype=np.float32)
                    probs = light_model.predict(X_light, batch_size=128, verbose=0).ravel()
                    for p, prob in zip(paths_for_pred, probs):
                        light_prob_by_path[p] = float(prob)
                else:
                    print("경고: Light 모델 추론에 사용할 유효한 샘플이 없습니다.")
    else:
        print(f"경고: 7. light_models 아래에 checkpoints/best.keras run 이 없어 Light 모델 추론을 생략합니다.")

    # 6) 측정 단위(한 전주/샘플)당 1행으로 정리
    #    각 행에는 x/y/z 축에 대한 모든 박스 정보를 문자열로 묶어서 넣는다.
    per_sample_rows = []
    axes = ["x", "y", "z"]
    for _, r in df.iterrows():
        base = {
            "server": r.get("server"),
            "project": r.get("project"),
            "pole_id": r.get("pole_id"),
            "sample_id": r.get("sample_id"),
            "csv_path": r.get("csv_path"),
            "label": r.get("label"),
            "label_text": r.get("label_text"),
        }
        label_val = int(r.get("label", 0)) if pd.notna(r.get("label")) else 0
        label_dir_name = "break" if label_val == 1 else "normal"
        # 13. evaluate_hard_model_2nd.py 에서 pred_boxes.json 은
        # <eval_dir>/<scope>/info/<break|normal> 아래에 저장된다.
        boxes_dir = scope_dir / "info" / label_dir_name
        boxes_json_path = boxes_dir / f"{base['sample_id']}_pred_boxes.json"
        if not boxes_json_path.exists():
            # 해당 샘플의 박스 JSON이 없으면 스킵
            continue
        try:
            with open(boxes_json_path, "r", encoding="utf-8") as f:
                boxes_json = json.load(f)
        except Exception:
            continue

        axes_dict = boxes_json.get("axes", {})
        axis_texts = {}
        for axis in axes:
            entries = axes_dict.get(axis, [])
            parts = []
            for entry in entries:
                parts.append(
                    "rank={rank},score={score},deg=[{dmin},{dmax}],h=[{hmin},{hmax}],"
                    "iou={iou},inside={inside},below_min={below}".format(
                        rank=entry.get("rank"),
                        score=entry.get("score"),
                        dmin=entry.get("degree_min"),
                        dmax=entry.get("degree_max"),
                        hmin=entry.get("height_min"),
                        hmax=entry.get("height_max"),
                        iou=entry.get("iou"),
                        inside=entry.get("inside_data_extent"),
                        below=entry.get("below_min_size"),
                    )
                )
            axis_texts[axis] = " | ".join(parts)

        # conf(score) 상위 3개 박스만 위치 좌표로 수집
        all_boxes = []
        for axis in axes:
            for entry in axes_dict.get(axis, []):
                score = entry.get("score")
                if score is None:
                    continue
                try:
                    s = float(score)
                except (TypeError, ValueError):
                    continue
                all_boxes.append({
                    "axis": axis,
                    "score": s,
                    "deg_min": entry.get("degree_min"),
                    "deg_max": entry.get("degree_max"),
                    "h_min": entry.get("height_min"),
                    "h_max": entry.get("height_max"),
                })
        all_boxes.sort(key=lambda b: b["score"], reverse=True)
        top3 = all_boxes[:3]
        top3_strs = []
        for i, b in enumerate(top3, start=1):
            top3_strs.append(
                f"{i}: axis={b['axis']}, score={b['score']:.3f}, "
                f"deg=[{b['deg_min']},{b['deg_max']}], h=[{b['h_min']},{b['h_max']}]"
            )

        row = base.copy()
        # Light 모델 기반 전주 단위 파단 확률/예측 라벨 추가
        lp = light_prob_by_path.get(str(base["csv_path"]))
        if lp is not None:
            row["light_prob_break"] = lp
            row["light_pred_break"] = int(lp >= 0.5)
        # 축별 박스 정보 문자열 컬럼
        for axis in axes:
            row[f"{axis}_boxes"] = axis_texts.get(axis, "")
        # conf(score) 상위 3개 박스 위치 좌표
        row["conf_top3_위치"] = " | ".join(top3_strs) if top3_strs else ""
        per_sample_rows.append(row)

    if not per_sample_rows:
        raise RuntimeError("측정 단위별 박스를 추출할 수 있는 데이터가 없습니다.")

    df_out = pd.DataFrame(per_sample_rows)

    # 6-1) 8번=약한 검출, 5번=강한 검출 (eval_recall_precision_labels.txt 기준)
    #      8번: [0] LIGHT t=0.80 (높은 기준 → 적게 울림), 5번: [0] LIGHT t=0.50 (낮은 기준 → 많이 울림)
    WEAK_THRESH = 0.80   # 8번 규칙
    STRONG_THRESH = 0.50  # 5번 규칙
    prob = df_out["light_prob_break"].astype(float).fillna(0.0)
    df_out["약한_검출"] = (prob >= WEAK_THRESH).astype(int)   # 8번
    df_out["강한_검출"] = (prob >= STRONG_THRESH).astype(int)  # 5번

    # 측정번호 = mesonu (같은 전주에 여러 번 측정 시 3, 4, 5, 6 ...)
    # 경로 예: 4. merge_data/normal/강원원주-202112/4114X505/4114X505_3_OUT_processed.csv -> 측정번호 3
    def _mesonu_from_sample_id(sid):
        if not sid or str(sid).strip() == "nan":
            return ""
        parts = str(sid).strip().split("_")
        if len(parts) >= 2 and parts[-1].isdigit():
            return parts[-1]
        return str(sid)

    def _mesonu_from_csv_path(csv_path):
        if not csv_path or str(csv_path).strip() == "nan":
            return ""
        stem = Path(csv_path).stem  # 4114X505_3_OUT_processed
        sid = stem.replace("_OUT_processed", "").strip()  # 4114X505_3
        return _mesonu_from_sample_id(sid)

    # sample_id(예: 4114X505_3)에서 측정번호 추출, 없거나 비면 csv_path 파일명에서 추출
    if "sample_id" in df_out.columns:
        df_out["측정번호"] = df_out["sample_id"].astype(str).apply(_mesonu_from_sample_id)
    else:
        df_out["측정번호"] = ""
    if "csv_path" in df_out.columns:
        empty = (df_out["측정번호"].astype(str).str.strip() == "") | (df_out["측정번호"].astype(str) == "nan")
        if empty.any():
            df_out.loc[empty, "측정번호"] = df_out.loc[empty, "csv_path"].astype(str).apply(_mesonu_from_csv_path)

    # 저장 시 제외할 컬럼 (label, 측정번호는 유지)
    drop_cols = ["sample_id", "csv_path", "light_pred_break"]
    df_out = df_out.drop(columns=[c for c in drop_cols if c in df_out.columns], errors="ignore")

    # 7) weight_test/export_2nd_eval_to_excel 아래에 엑셀로 저장 (각 행 = 한 전주/샘플)
    out_dir = current_dir / "export_2nd_eval_to_excel"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "hard_2nd_merge_data_predictions.xlsx"

    df_out.to_excel(out_path, index=False)

    # 8) 조건부 색상: 8번(약한_검출) 충족 → 빨간색(파단), 5번만 충족(8번 미충족) → 주황색(의심)
    try:
        wb = load_workbook(out_path)
        ws = wb.active

        header = {cell.value: cell.column for cell in ws[1] if cell.value}
        col_weak = header.get("약한_검출")
        col_strong = header.get("강한_검출")

        red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")   # 파단(8번)
        orange_fill = PatternFill(start_color="FFFFCC99", end_color="FFFFCC99", fill_type="solid")  # 5번~8번 사이(의심)

        if col_weak is not None and col_strong is not None:
            for row_idx in range(2, ws.max_row + 1):
                v_weak = ws.cell(row=row_idx, column=col_weak).value
                v_strong = ws.cell(row=row_idx, column=col_strong).value
                try:
                    weak_ok = int(v_weak) == 1 if v_weak is not None else False
                    strong_ok = int(v_strong) == 1 if v_strong is not None else False
                except (TypeError, ValueError):
                    weak_ok, strong_ok = False, False
                fill = None
                if weak_ok:
                    fill = red_fill   # 8번 포함 → 파단(빨간색)
                elif strong_ok:
                    fill = orange_fill  # 5번만 포함 → 주황색(의심)
                if fill is not None:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill

        wb.save(out_path)
        print(f"저장 및 강조 표시 완료: {out_path} (8번=빨간색, 5번~8번 사이=주황색)")
    except Exception as e:
        print(f"경고: 엑셀 강조 표시 적용 중 오류 발생: {e}")


if __name__ == "__main__":
    main()

